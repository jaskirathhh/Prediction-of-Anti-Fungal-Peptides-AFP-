import pandas as pd
import xlwt
import xlrd
import openpyxl as xl
from xlutils.copy import copy
from modlamp.descriptors import PeptideDescriptor, GlobalDescriptor
import sklearn.externals.joblib
import openpyxl as xl
from openpyxl import Workbook
  
########################################################TRAINING DATA##############################################################


trainingData = pd.read_csv(r"train.csv")                            #reading CSV training data
trainingData.to_excel(r"train.xlsx", index = None, header=True)     #converting CSV to Excel

wb = Workbook()                 				    #creating workbook
wb.save("output.xlsx")						    #saving newly created empty workbook
readingBook  = xl.load_workbook("train.xlsx") 			    #openning training Data book
readingSheet = readingBook.worksheets[0]			    #openning training Data sheet

outputFile ="output.xlsx"					    #openning newly created output file
writingBook = xl.load_workbook(outputFile) 
writingSheet = writingBook.active

rows = readingSheet.max_row 					    #no of rows
cols = readingSheet.max_column					    #no of cols

for i in range (1, rows+1): 					    # reading values from training Data and writing in output file
    for j in range (1, cols+1): 									
        cellValue = readingSheet.cell(row = i, column = j)					  
        writingSheet.cell(row = i, column = j).value = cellValue.value
            
  

"""Rewriting label values to output file to get rid of some issues with libraries used """


labelArray = []
for i in range (2, rows+1):
    if (readingSheet.cell(row = i, column = 2).value=="-1"):
        labelArray.append(-1)
    else:
        labelArray.append(1)

for i in range (2, rows+1):
    writingSheet.cell(row = i,column = 2).value = labelArray[i-2]



"""Using Types of amino acid  present in sequence as a feature to determine AFP sequence"""

aminoAcid = ["(G)Glycine","(P)Proline","(A)Alanine","(V)Valine","(L)Leucine","(I)Isoleucine","(M)Methionine","(C)Cysteine","(F)Phenylalanine",
"(Y)Tyrosine","(W)Tryptophan","(H)Histidine","(K)Lysine","(R)Arginine","(Q)Glutamine","(N)Asparagine","(E)Glutamic Acid",
"(D)Aspartic Acid","(S)Serine","(T)Threonine"]


lengthAminoAcid = len(aminoAcid)				    #total features each corresponding to each amino acid

"""Function to find amino acidcomposion """

def countResidue(sequence,residue):
    count = 0
    for ele in sequence:
        if ele == residue:
            count+=1
    return count

filler = 0							    #writing feature names in excel sheet
for i in range(cols+1,cols+len(aminoAcid)+1):
    writingSheet.cell(row = 1, column = i).value = aminoAcid[filler]
    filler+=1
    


for i in range(2,rows+1):					    #filling amino acid composition of each type of amino acid in sequence
    for j in range(cols+1,cols+len(aminoAcid)+1):
        ele = aminoAcid[j-cols-1][1]
        pepSequence = readingSheet.cell(row = i,column = cols).value
        writingSheet.cell(row = i,column = j).value = float((countResidue(pepSequence,ele)/len(pepSequence))*100)


"""Some more features other than amino acid composition of each amino acid  in sequence"""

newFeatures = ['MW', 'ChargeDensity', 'pI',
               'InstabilityInd', 'Aromaticity', 'AliphaticInd',
               'BomanInd', 'HydRatio']


               							    #writing feature names in excel sheet
for i in range(cols+len(aminoAcid)+1,cols+len(aminoAcid)+len(newFeatures)+1):
    writingSheet.cell(row = 1, column = i).value = newFeatures[i-(cols+len(aminoAcid)+1)]



for i in range(2,rows+1):					    #filling feature value in excel sheet  
    pepSequencee = readingSheet.cell(row = i,column = cols).value
    desc = GlobalDescriptor(pepSequencee)
    desc.calculate_all(amide=True)
    array = desc.descriptor.tolist()
    countt = 1
    for j in range(cols+len(aminoAcid)+1,cols+len(aminoAcid)+1+len(newFeatures)):
        writingSheet.cell(row = i,column = j).value = float(array[0][countt])
        countt+=1
        


writingBook.save(str(outputFile))				    #saving all data to output file


##################################################################TESTING DATA####################################################


trainingData = pd.read_csv(r"test.csv")                             #reading CSV training data
trainingData.to_excel(r"test.xlsx", index = None, header=True)      #converting CSV to Excel

wb = Workbook()                 				    #creating workbook for processed test data 
wb.save("testoutput.xlsx")					    #saving new workbook


readingBook  = xl.load_workbook("test.xlsx") 			    #openning training Data book
readingSheet = readingBook.worksheets[0]			    #openning training Data sheet

outputFile ="testoutput.xlsx"					    #opening  newly created workbook for processed data
writingBook = xl.load_workbook(outputFile) 
writingSheet = writingBook.active

rows = readingSheet.max_row 					    #no of rows
cols = readingSheet.max_column					    #no of cols


for i in range (1, rows+1): 					    # creating processed test data 
    for j in range (1, cols+1): 
        cellValue = readingSheet.cell(row = i, column = j)					        
        writingSheet.cell(row = i, column = j).value = cellValue.value
            
  

"""Same as above"""
filler = 0
for i in range(cols+1,cols+len(aminoAcid)+1):
    writingSheet.cell(row = 1, column = i).value = aminoAcid[filler]
    filler+=1    

for i in range(2,rows+1):
    for j in range(cols+1,cols+len(aminoAcid)+1):
        ele = aminoAcid[j-cols-1][1]
        pepSequence = readingSheet.cell(row = i,column = cols).value
        writingSheet.cell(row = i,column = j).value = float((countResidue(pepSequence,ele)/len(pepSequence))*100)


"""Same as above"""
for i in range(cols+len(aminoAcid)+1,cols+len(aminoAcid)+1+len(newFeatures)):
    writingSheet.cell(row = 1, column = i).value = newFeatures[i-(cols+len(aminoAcid)+1)]

for i in range(2,rows+1):
    pepSequencee = readingSheet.cell(row = i,column = cols).value
    desc = GlobalDescriptor(pepSequencee)
    desc.calculate_all(amide=True)
    array = desc.descriptor.tolist()
    countt = 1
    for j in range(cols+len(aminoAcid)+1,cols+len(aminoAcid)+1+len(newFeatures)):
        writingSheet.cell(row = i,column = j).value = float(array[0][countt])
        countt+=1
        





    
writingBook.save(str(outputFile))					#saving processed output data



#####################################################################PREDICTION#####################################################

finalread = pd.read_excel(r"output.xlsx")							 
finalread.to_csv(r"trainingFeature.csv")
dataset = pd.read_csv("trainingFeature.csv")
X = dataset.iloc[:,4:32].values						#selecting features from Training Data
Y = dataset.iloc[:,2].values						#selecting result column for Training





finalread1 = pd.read_excel(r"testoutput.xlsx")
finalread1.to_csv(r"testsdf.csv")
dataset1 = pd.read_csv("testsdf.csv")
X_calculated = dataset1.iloc[:,3:31].values				#selecting features calculated using test data





from xgboost import XGBClassifier					#trainging our model from training data					
classifier = XGBClassifier()
classifier.fit(X,Y)

y_pred =classifier.predict(X_calculated)				#storing prediction of our model




dataset2 = pd.read_csv("sample.csv")
result = y_pred.tolist()						#saving prediction result to PREDICTION.csv
idarray = dataset2.iloc[:,0].values
idarray = idarray.tolist()
dict  = {"ID":idarray,"Label":result}
df = pd.DataFrame(dict)
  
df.to_csv('PREDICTION.csv',index = False) 


